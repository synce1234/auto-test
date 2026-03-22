#!/usr/bin/env python3
"""
Tạo file Excel template chứa danh sách test cases.
Chạy 1 lần để khởi tạo: python3 create_template.py
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.path.join(os.path.dirname(__file__), "test_cases.xlsx")

# ─── Màu sắc ──────────────────────────────────────────────────────────────────
COLOR_HEADER_BG   = "1F4E79"   # xanh đậm
COLOR_HEADER_FONT = "FFFFFF"   # trắng
COLOR_SECTION_BG  = "2E75B6"   # xanh section
COLOR_ALT_ROW     = "DEEAF1"   # xanh nhạt xen kẽ
COLOR_PASS        = "C6EFCE"   # xanh lá nhạt
COLOR_FAIL        = "FFC7CE"   # đỏ nhạt
COLOR_SKIP        = "FFEB9C"   # vàng nhạt
COLOR_NOT_RUN     = "F2F2F2"   # xám nhạt


def header_style():
    return {
        "fill": PatternFill("solid", fgColor=COLOR_HEADER_BG),
        "font": Font(bold=True, color=COLOR_HEADER_FONT, size=11),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            bottom=Side(style="medium", color="FFFFFF"),
        ),
    }


def apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)


def create_template():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Test Cases ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Test Cases"

    # Header row
    columns = [
        ("Testcase ID",                         15),
        ("Phân Cấp",                            12),
        ("Nội Dung Test\n(Test Title)",         35),
        ("Điều Kiện / Dữ Liệu Test\n(Pre-condition / Test Data)", 35),
        ("Các Bước Thực Hiện\n(Test Steps)",    45),
        ("Kết Quả Mong Đợi\n(Expected Results)",35),
        ("Kết Quả Thực Tế\n(Actual Results)",   35),
        ("Trạng Thái\n(Status)",                12),
        ("Thời Gian Chạy\n(Duration)",          15),
        ("Ghi Chú\n(Notes)",                    30),
    ]

    h_style = header_style()
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        apply_style(cell, h_style)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 45
    ws.freeze_panes = "A2"

    # Data validation dropdown cho cột Status (col 8)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_status = DataValidation(
        type="list",
        formula1='"PASS,FAIL,SKIP,NOT RUN"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_status)
    dv_status.sqref = "H2:H1000"

    # Data validation dropdown cho cột Phân Cấp (col 2)
    dv_priority = DataValidation(
        type="list",
        formula1='"P0 - Critical,P1 - High,P2 - Medium,P3 - Low"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv_priority)
    dv_priority.sqref = "B2:B1000"

    # Ví dụ 1 row mẫu
    sample = [
        "TC_001",
        "P1 - High",
        "Mở file PDF từ danh sách",
        "App đã cài, có ít nhất 1 file PDF trong danh sách",
        "1. Mở app\n2. Xem danh sách file\n3. Click vào file PDF đầu tiên",
        "PDF viewer mở thành công, hiển thị nội dung file",
        "",
        "NOT RUN",
        "",
        "",
    ]
    for col_idx, value in enumerate(sample, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
    ws.row_dimensions[2].height = 60

    # ── Sheet 2: Summary ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    summary_headers = [
        ("Thống Kê", "Số Lượng"),
    ]
    summary_rows = [
        ("Tổng Test Cases",   '=COUNTA(\'Test Cases\'!A2:A1000)'),
        ("PASS",              '=COUNTIF(\'Test Cases\'!H2:H1000,"PASS")'),
        ("FAIL",              '=COUNTIF(\'Test Cases\'!H2:H1000,"FAIL")'),
        ("SKIP",              '=COUNTIF(\'Test Cases\'!H2:H1000,"SKIP")'),
        ("NOT RUN",           '=COUNTIF(\'Test Cases\'!H2:H1000,"NOT RUN")'),
        ("Tỉ Lệ Pass (%)",   '=IF(B2=0,0,ROUND(B3/B2*100,1))'),
    ]

    # Header
    for col_idx, val in enumerate(["Thống Kê", "Số Lượng"], start=1):
        cell = ws2.cell(row=1, column=col_idx, value=val)
        apply_style(cell, h_style)

    # Data rows
    fills = {
        "Tổng Test Cases":  "1F4E79",
        "PASS":             COLOR_PASS,
        "FAIL":             COLOR_FAIL,
        "SKIP":             COLOR_SKIP,
        "NOT RUN":          COLOR_NOT_RUN,
        "Tỉ Lệ Pass (%)":  "E2EFDA",
    }
    font_colors = {
        "Tổng Test Cases": "FFFFFF",
    }

    for row_idx, (label, formula) in enumerate(summary_rows, start=2):
        c_label = ws2.cell(row=row_idx, column=1, value=label)
        c_value = ws2.cell(row=row_idx, column=2, value=formula)
        bg = fills.get(label, "FFFFFF")
        fc = font_colors.get(label, "000000")
        for c in [c_label, c_value]:
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(bold=(label == "Tổng Test Cases" or label == "Tỉ Lệ Pass (%)"),
                          color=fc)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        ws2.row_dimensions[row_idx].height = 22

    wb.save(OUTPUT)
    print(f"✅ Đã tạo: {OUTPUT}")
    print(f"   Sheet 1: 'Test Cases' — điền test cases vào đây")
    print(f"   Sheet 2: 'Summary'    — tự động tính thống kê")


if __name__ == "__main__":
    create_template()
