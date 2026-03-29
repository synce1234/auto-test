"""
TC Manager - Đọc/ghi test case info từ Excel, generate report sau khi chạy test.

Cách dùng:
  from test_cases.tc_manager import TCManager
  tc = TCManager()
  info = tc.get("TC_001")          # lấy info của 1 TC
  tc.update_result("TC_001", "PASS", duration=2.3, actual="Mở thành công")
  tc.save_report("reports/run_001.xlsx")
"""
import os
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "test_cases.xlsx")

# Vị trí các cột trong sheet "Test Cases" (1-based)
COL = {
    "id":           1,
    "priority":     2,
    "title":        3,
    "precondition": 4,
    "steps":        5,
    "expected":     6,
    "actual":       7,
    "status":       8,
    "duration":     9,
    "notes":        10,
}

STATUS_COLORS = {
    "PASS":         "C6EFCE",
    "FAIL":         "FFC7CE",
    "SKIP":         "FFEB9C",
    "NOT RUN":      "F2F2F2",
    "NEED CONFIRM": "BDD7EE",
}


class TCManager:
    def __init__(self, excel_path: str = EXCEL_PATH):
        self.excel_path = excel_path
        self._cases: dict[str, dict] = {}
        self._results: dict[str, dict] = {}
        self._load()

    def _load(self):
        """Đọc tất cả test cases từ Excel."""
        if not os.path.exists(self.excel_path):
            print(f"[WARN] Không tìm thấy {self.excel_path}")
            print("       Chạy: python3 test_cases/create_template.py")
            return

        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb["Test Cases"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            tc_id = row[COL["id"] - 1]
            if not tc_id:
                continue
            self._cases[str(tc_id)] = {
                "id":           str(tc_id),
                "priority":     row[COL["priority"] - 1] or "",
                "title":        row[COL["title"] - 1] or "",
                "precondition": row[COL["precondition"] - 1] or "",
                "steps":        row[COL["steps"] - 1] or "",
                "expected":     row[COL["expected"] - 1] or "",
                "status":       row[COL["status"] - 1] or "NOT RUN",
            }

        print(f"[TC] Đã load {len(self._cases)} test cases từ {os.path.basename(self.excel_path)}")

    def get(self, tc_id: str) -> dict:
        """Lấy thông tin của 1 test case."""
        return self._cases.get(tc_id, {})

    def get_all(self) -> dict:
        return self._cases

    def update_result(
        self,
        tc_id: str,
        status: str,
        actual: str = "",
        duration: float = 0.0,
        notes: str = "",
    ):
        """Cập nhật kết quả test cho 1 TC."""
        self._results[tc_id] = {
            "status":   status.upper(),
            "actual":   actual,
            "duration": f"{duration:.2f}s" if duration else "",
            "notes":    notes,
        }

    def save_report(self, report_path: str = None):
        """
        Ghi kết quả vào file Excel report riêng (không sửa file gốc).
        report_path: đường dẫn file output, mặc định tạo theo timestamp.
        """
        if not report_path:
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            report_path = os.path.join(
                os.path.dirname(__file__), "reports", f"report_{ts}.xlsx"
            )

        os.makedirs(os.path.dirname(report_path), exist_ok=True)

        # Mở file gốc làm base
        if os.path.exists(self.excel_path):
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb["Test Cases"]
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Cases"

        # Ghi kết quả vào các cột Actual/Status/Duration/Notes
        for row in ws.iter_rows(min_row=2):
            tc_id_cell = row[COL["id"] - 1]
            tc_id = str(tc_id_cell.value or "")
            if not tc_id or tc_id not in self._results:
                continue

            result = self._results[tc_id]
            status = result["status"]

            row[COL["actual"] - 1].value   = result["actual"]
            row[COL["status"] - 1].value   = status
            row[COL["duration"] - 1].value = result["duration"]
            row[COL["notes"] - 1].value    = result["notes"]

            # Tô màu cả hàng theo status
            fill_color = STATUS_COLORS.get(status, "FFFFFF")
            for cell in row:
                if cell.value is not None or cell.column <= len(COL):
                    cell.fill = PatternFill("solid", fgColor=fill_color)

        # Thêm sheet Run Info
        if "Run Info" in wb.sheetnames:
            del wb["Run Info"]
        ws_info = wb.create_sheet("Run Info")
        ws_info["A1"] = "Thông tin lần chạy"
        ws_info["A1"].font = Font(bold=True, size=13)
        total        = len(self._results)
        passed       = sum(1 for r in self._results.values() if r["status"] == "PASS")
        failed       = sum(1 for r in self._results.values() if r["status"] == "FAIL")
        skipped      = sum(1 for r in self._results.values() if r["status"] == "SKIP")
        need_confirm = sum(1 for r in self._results.values() if r["status"] == "NEED CONFIRM")
        run_info = [
            ("Thời điểm chạy",  datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Tổng TC",         total),
            ("PASS",            passed),
            ("FAIL",            failed),
            ("SKIP",            skipped),
            ("NEED CONFIRM",    need_confirm),
        ]
        run_info.append(("Tỉ lệ Pass", f"{passed/total*100:.1f}%" if total else "N/A"))

        for row_idx, (label, value) in enumerate(run_info, start=2):
            ws_info.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws_info.cell(row=row_idx, column=2, value=value)
        ws_info.column_dimensions["A"].width = 20
        ws_info.column_dimensions["B"].width = 25

        wb.save(report_path)
        print(f"\n[REPORT] Đã lưu: {report_path}")
        self._print_summary()
        return report_path

    def _print_summary(self):
        total        = len(self._results)
        passed       = sum(1 for r in self._results.values() if r["status"] == "PASS")
        failed       = sum(1 for r in self._results.values() if r["status"] == "FAIL")
        skipped      = sum(1 for r in self._results.values() if r["status"] == "SKIP")
        need_confirm = sum(1 for r in self._results.values() if r["status"] == "NEED CONFIRM")
        rate         = f"{passed/total*100:.1f}%" if total else "N/A"
        print(f"  Tổng: {total} | ✅ PASS: {passed} | ❌ FAIL: {failed} | ⏭ SKIP: {skipped} | 🔍 NEED CONFIRM: {need_confirm} | Tỉ lệ: {rate}")
