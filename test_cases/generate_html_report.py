#!/usr/bin/env python3
"""
HTML Dashboard Report Generator.

Đọc kết quả từ TCManager Excel, tạo báo cáo HTML với:
  - Summary cards (Total / Pass / Fail / Skip / Pass rate)
  - Donut chart tổng quan
  - Lịch sử các lần chạy (click để cập nhật in-place)
  - Bảng TC có badge trạng thái, screenshot inline, link video

Dùng:
  python3 test_cases/generate_html_report.py [report.xlsx] [-o output.html]
"""
import os
import sys
import base64
import datetime
import argparse
import json
import openpyxl

ROOT        = os.path.join(os.path.dirname(__file__), "..")
TC_EXCEL    = os.path.join(os.path.dirname(__file__), "test_cases.xlsx")
REPORTS_DIR = os.path.join(ROOT, "reports")

# ─── Data loading ──────────────────────────────────────────────────────────────

def _load_tc_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path)
    ws = wb["Test Cases"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tc_id = row[0]
        if not tc_id:
            continue
        rows.append({
            "id":           str(tc_id),
            "group":        str(row[1] or "").strip(),
            "title":        str(row[2] or "").strip(),
            "precondition": str(row[3] or "").strip(),
            "steps":        str(row[4] or "").strip(),
            "expected":     str(row[5] or "").strip(),
            "actual":       str(row[6] or "").strip(),
            "status":       str(row[7] or "NOT RUN").strip().upper(),
            "duration":     str(row[8] or "").strip(),
            "notes":        str(row[9] or "").strip(),
        })
    return rows


def _asset_dirs(run_ts: str | None) -> tuple[str, str]:
    if run_ts:
        ss  = os.path.join(REPORTS_DIR, "screenshots", run_ts)
        vid = os.path.join(REPORTS_DIR, "videos",      run_ts)
    else:
        ss  = os.path.join(REPORTS_DIR, "screenshots")
        vid = os.path.join(REPORTS_DIR, "videos")
    return ss, vid


def _log_dir(run_ts: str | None) -> str:
    return os.path.join(REPORTS_DIR, "logs", run_ts) if run_ts else os.path.join(REPORTS_DIR, "logs")


def _find_all_logs(tc_id: str, run_ts: str | None = None) -> list[str]:
    return _find_all_in_dir(_log_dir(run_ts), tc_id, ".txt")


def _find_all_in_dir(directory: str, tc_id: str, ext: str) -> list[str]:
    """Trả về danh sách tất cả file khớp TC ID trong directory.

    Ưu tiên match theo TC ID (tc_sm_001, tc_pdf_003, ...) trong tên file.
    Fallback: match theo số thứ tự 3 chữ số (tc_001, tc001) cho TC cũ.
    """
    if not os.path.isdir(directory):
        return []

    # Pattern mới: tc_id trực tiếp trong tên file (vd: tc_sm_001, tc_pdf_003, tc_dm_007)
    pat_direct = tc_id.lower()  # vd: "tc_sm_001"

    # Pattern cũ (backward compat): chỉ lấy phần số cuối
    import re as _re
    num_match = _re.search(r'(\d+)$', tc_id)
    tc_num = num_match.group(1).lstrip("0") if num_match else ""
    pat_numeric1 = f"tc_{tc_num.zfill(3)}" if tc_num else ""  # tc_001
    pat_numeric2 = f"tc{tc_num.zfill(3)}"  if tc_num else ""  # tc001

    results = []
    for fname in sorted(os.listdir(directory)):
        if not fname.lower().endswith(ext):
            continue
        base = fname[: -len(ext)].lower()
        if (pat_direct and pat_direct in base) or \
           (pat_numeric1 and pat_numeric1 in base) or \
           (pat_numeric2 and pat_numeric2 in base):
            results.append(os.path.join(directory, fname))
    return results


def _find_all_screenshots(tc_id: str, run_ts: str | None = None) -> list[str]:
    ss_dir, _ = _asset_dirs(run_ts)
    return _find_all_in_dir(ss_dir, tc_id, ".png")


def _find_all_videos(tc_id: str, run_ts: str | None = None) -> list[str]:
    _, vid_dir = _asset_dirs(run_ts)
    return _find_all_in_dir(vid_dir, tc_id, ".mp4")


def _img_to_b64(path: str) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()


def _has_test_cases_sheet(path: str) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        return "Test Cases" in wb.sheetnames
    except Exception:
        return False


def _list_all_in_dir(directory: str, ext: str) -> list[str]:
    """Liệt kê tất cả file có đuôi ext trong directory."""
    if not os.path.isdir(directory):
        return []
    return sorted(os.path.join(directory, f)
                  for f in os.listdir(directory)
                  if f.lower().endswith(ext))


def _attach_assets(cases: list[dict], ts_str: str | None) -> list[dict]:
    """
    Thêm screenshots (base64) và videos (relative path) vào mỗi case.
    - Run mới (có TC prefix): gắn đúng file vào từng TC.
    - Run cũ (không có TC prefix): gắn tất cả file vào TC đầu tiên.
    """
    import re as _re
    ss_dir, vid_dir = _asset_dirs(ts_str)

    # Kiểm tra run có dùng naming mới (TC_001_xxx) chưa
    _tc_pat = _re.compile(r'tc[_a-z]*\d{3}', _re.IGNORECASE)
    all_ss  = _list_all_in_dir(ss_dir,  ".png")
    all_vid = _list_all_in_dir(vid_dir, ".mp4")
    has_prefix = any(_tc_pat.search(os.path.basename(p)) for p in all_ss + all_vid)

    all_log = _list_all_in_dir(_log_dir(ts_str), ".txt")

    result = []
    for i, c in enumerate(cases):
        if has_prefix:
            ss_paths  = _find_all_screenshots(c["id"], ts_str)
            vid_paths = _find_all_videos(c["id"], ts_str)
            log_paths = _find_all_logs(c["id"], ts_str)
        elif i == 0:
            ss_paths  = all_ss
            vid_paths = all_vid
            log_paths = all_log
        else:
            ss_paths, vid_paths, log_paths = [], [], []

        ss_b64 = []
        for p in ss_paths:
            if os.path.exists(p):
                try:
                    ss_b64.append(_img_to_b64(p))
                except Exception:
                    pass
        vids = []
        for p in vid_paths:
            if os.path.exists(p):
                vids.append(os.path.relpath(p, REPORTS_DIR))

        log_content = ""
        for p in log_paths:
            if os.path.exists(p):
                try:
                    with open(p, encoding="utf-8", errors="replace") as _f:
                        chunk = _f.read(10000)
                    if log_content:
                        log_content += "\n\n─── next log ───\n\n"
                    log_content += chunk
                except Exception:
                    pass

        result.append({**c, "screenshots": ss_b64, "videos": vids, "log": log_content})
    return result


def _load_history(current_ts: str | None = None) -> list[dict]:
    """Đọc tất cả report Excel (trừ run hiện tại), kèm cases + assets."""
    history = []
    if not os.path.isdir(REPORTS_DIR):
        return history
    for fname in sorted(os.listdir(REPORTS_DIR)):
        if not (fname.startswith("report_") and fname.endswith(".xlsx")):
            continue
        ts_str = fname.replace("report_", "").replace(".xlsx", "")
        if ts_str == current_ts:
            continue
        fpath = os.path.join(REPORTS_DIR, fname)
        try:
            rows   = _load_tc_excel(fpath) if _has_test_cases_sheet(fpath) else []
            ts     = datetime.datetime.strptime(ts_str, "%Y%m%d_%H%M%S")
            total  = len(rows)
            passed = sum(1 for r in rows if r["status"] == "PASS")
            failed = sum(1 for r in rows if r["status"] == "FAIL")
            skip   = sum(1 for r in rows if r["status"] == "SKIP")
            not_run = total - passed - failed - skip
            history.append({
                "ts_str":  ts_str,
                "ts":      ts.strftime("%d/%m/%Y %H:%M"),
                "total":   total,
                "passed":  passed,
                "failed":  failed,
                "skip":    skip,
                "not_run": not_run,
                "rate":    f"{passed/total*100:.0f}%" if total else "N/A",
                "cases":   _attach_assets(rows, ts_str),
            })
        except Exception:
            pass
    return history[-10:]


# ─── HTML helpers ───────────────────────────────────────────────────────────────

_STATUS_COLOR = {
    "PASS":         ("#C6EFCE", "#276221", "✅"),
    "FAIL":         ("#FFC7CE", "#9C0006", "❌"),
    "SKIP":         ("#FFEB9C", "#7D6608", "⏭"),
    "NOT RUN":      ("#F2F2F2", "#555555", "⬜"),
    "NEED CONFIRM": ("#BDD7EE", "#1F497D", "👁"),
}


def _escape(text: str) -> str:
    return (str(text or "")
            .replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _nl2br(text: str) -> str:
    return _escape(text).replace("\n", "<br>") if text else "—"


def _history_section(history: list[dict]) -> str:
    if not history:
        return ""
    rows = ""
    for h in reversed(history):
        bar_w  = int(h["passed"] / h["total"] * 100) if h["total"] else 0
        color  = "#276221" if bar_w == 100 else ("#9C0006" if bar_w < 50 else "#7D6608")
        fail_c = "color:#9C0006;font-weight:bold" if h["failed"] else ""
        rows += f"""
        <tr class="hist-row" id="hist-row-{h['ts_str']}"
            onclick="loadRun('{h['ts_str']}')"
            style="cursor:pointer" title="Click để xem kết quả lần chạy này">
          <td style="color:#2b6cb0;font-weight:500">🔍 {h['ts']}</td>
          <td class="center">{h['total']}</td>
          <td class="center" style="color:#276221;font-weight:bold">{h['passed']}</td>
          <td class="center" style="{fail_c}">{h['failed']}</td>
          <td class="center">{h['skip']}</td>
          <td>
            <div style="background:#eee;border-radius:4px;height:18px;width:100%">
              <div style="background:{color};width:{bar_w}%;height:100%;border-radius:4px;
                          display:flex;align-items:center;justify-content:center;
                          font-size:11px;color:#fff;font-weight:bold;min-width:28px">
                {h['rate']}
              </div>
            </div>
          </td>
        </tr>"""
    return f"""<h3 style="font-size:14px;font-weight:600;margin-bottom:12px;color:var(--muted)">
      📈 Lịch sử chạy gần nhất
    </h3>
    <div style="overflow-x:auto">
      <table class="history-table" style="width:100%">
        <thead>
          <tr>
            <th>Thời điểm (click xem)</th><th>Tổng</th>
            <th>Pass</th><th>Fail</th><th>Skip</th><th style="min-width:120px">Pass Rate</th>
          </tr>
        </thead>
        <tbody>{rows}</tbody>
      </table>
    </div>"""


def _tc_row(c: dict) -> str:
    """Python-render một TC row cho initial load (current run)."""
    status = c["status"]
    bg, fg, icon = _STATUS_COLOR.get(status, ("#eee", "#333", "?"))
    row_cls = "row-fail" if status == "FAIL" else ("row-pass" if status == "PASS" else "")
    badge   = f'<span class="badge" style="background:{bg};color:{fg}">{icon} {status}</span>'
    actual_color = "#276221" if status == "PASS" else ("#9C0006" if status == "FAIL" else "inherit")

    # Multiple screenshots
    tc_id_esc = _escape(c["id"])
    ss_cells = ""
    for b64 in c.get("screenshots", []):
        ss_cells += (f'<img src="data:image/png;base64,{b64}" '
                     f'style="height:44px;border-radius:4px;border:1px solid #ddd;'
                     f'cursor:zoom-in;margin:2px" '
                     f"onclick=\"showSS('{tc_id_esc}',this.src)\" "
                     f'title="Xem screenshot">')
    ss_cells = ss_cells or "—"

    # Multiple videos
    vid_cells = ""
    for rel in c.get("videos", []):
        vid_cells += f'<a href="{rel}" target="_blank" style="font-size:20px;margin:2px" title="Xem video">🎬</a>'
    vid_cells = vid_cells or "—"

    # Console log
    log_content = c.get("log", "")
    if log_content:
        log_cell = (
            f'<details><summary style="cursor:pointer;font-size:12px;color:#2b6cb0;white-space:nowrap">📋 Log</summary>'
            f'<pre style="font-size:11px;white-space:pre-wrap;max-height:300px;overflow:auto;'
            f'background:#1e1e1e;color:#d4d4d4;padding:8px;border-radius:4px;margin-top:4px;text-align:left">'
            f'{_escape(log_content)}</pre></details>'
        )
    else:
        log_cell = "—"

    return f"""
    <tr class="{row_cls}">
      <td><strong>{_escape(c['id'])}</strong></td>
      <td class="group-cell">{_escape(c['group'])}</td>
      <td>{_nl2br(c['title'])}</td>
      <td style="min-width:180px">{_nl2br(c['steps'])}</td>
      <td style="min-width:180px">{_nl2br(c['expected'])}</td>
      <td style="min-width:180px;color:{actual_color}">{_nl2br(c['actual']) or '—'}</td>
      <td class="center">{badge}</td>
      <td class="center" style="white-space:nowrap">{_escape(c['duration'])}</td>
      <td class="center">{ss_cells}</td>
      <td class="center">{vid_cells}</td>
      <td class="center">{log_cell}</td>
    </tr>"""


def _group_section(group_name: str, cases: list[dict]) -> str:
    total  = len(cases)
    passed = sum(1 for c in cases if c["status"] == "PASS")
    rows   = "".join(_tc_row(c) for c in cases)
    return f"""
  <div class="section">
    <div class="section-header">
      <h2>🗂 {_escape(group_name)}</h2>
      <span class="group-badge">{passed}/{total} Pass</span>
    </div>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th>TC ID</th><th>Nhóm</th><th>Nội dung</th>
            <th>Các bước</th><th>Kết quả mong đợi</th><th>Kết quả thực tế</th>
            <th>Trạng thái</th><th>Thời gian</th><th>Screenshot</th><th>Video</th><th>Log</th>
          </tr>
        </thead>
        <tbody>{rows}</tbody>
      </table>
    </div>
  </div>"""


# ─── Main HTML builder ─────────────────────────────────────────────────────────

def generate_html(cases: list[dict], run_info: dict, history: list[dict],
                  run_ts: str | None = None) -> str:
    # Attach assets to current run cases, chỉ giữ lại các TC đã chạy
    current_cases = [c for c in _attach_assets(cases, run_ts) if c["status"] != "NOT RUN"]

    total   = len(current_cases)
    passed  = sum(1 for c in current_cases if c["status"] == "PASS")
    failed  = sum(1 for c in current_cases if c["status"] == "FAIL")
    skipped = sum(1 for c in current_cases if c["status"] == "SKIP")
    not_run = 0
    rate    = f"{passed/total*100:.1f}%" if total else "N/A"

    run_display = run_info.get("timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    run_src     = run_info.get("source", "Manual")

    log_dir_path = os.path.join(REPORTS_DIR, "logs", run_ts) if run_ts else None
    has_logs     = bool(log_dir_path and os.path.isdir(log_dir_path) and os.listdir(log_dir_path))
    log_dl_btn   = (
        f'&nbsp;·&nbsp;<a href="/api/logs/{run_ts}/download" download="logs_{run_ts}.zip" '
        f'style="color:#93c5fd;font-size:12px;text-decoration:none;opacity:.85;">📥 Download Logs</a>'
    ) if has_logs else ""

    # TC tables for initial load (current run)
    groups: dict[str, list] = {}
    for c in current_cases:
        g = c["group"] or "Uncategorized"
        groups.setdefault(g, []).append(c)
    tc_sections_html = "".join(_group_section(g, cs) for g, cs in groups.items())

    # Build RUNS_DATA JSON (embed all run data for JS)
    current_run_data = {
        "ts":      run_display,
        "total":   total,
        "passed":  passed,
        "failed":  failed,
        "skip":    skipped,
        "not_run": not_run,
        "rate":    rate,
        "cases":   current_cases,
    }
    hist_runs: dict = {}
    for h in history:
        hist_runs[h["ts_str"]] = {
            "ts":      h["ts"],
            "total":   h["total"],
            "passed":  h["passed"],
            "failed":  h["failed"],
            "skip":    h["skip"],
            "not_run": h.get("not_run", 0),
            "rate":    h["rate"],
            "cases":   h["cases"],
        }
    runs_data_json = json.dumps({"current": current_run_data, **hist_runs},
                                ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>PDF Reader — Auto Test Report</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --bg: #f5f7fa; --card: #fff; --border: #e2e8f0;
    --text: #1a202c; --muted: #718096;
    --pass: #276221; --fail: #9C0006; --skip: #7D6608;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
          background: var(--bg); color: var(--text); font-size: 14px; }}
  .header {{ background: linear-gradient(135deg,#1a365d,#2b6cb0);
             color: #fff; padding: 24px 32px; }}
  .header h1 {{ font-size: 22px; font-weight: 700; margin-bottom: 6px; }}
  .header .meta {{ font-size: 13px; opacity: 0.8; }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 24px 20px; }}
  /* Run banner */
  .run-banner {{ background: #ebf8ff; border: 1px solid #bee3f8; border-radius: 8px;
                 padding: 10px 16px; margin-bottom: 20px; font-size: 13px;
                 display: flex; align-items: center; gap: 12px; }}
  .run-banner a {{ color: #2b6cb0; font-weight: 600; text-decoration: none; }}
  .run-banner a:hover {{ text-decoration: underline; }}
  /* Cards */
  .cards {{ display: grid; grid-template-columns: repeat(5,1fr); gap: 16px; margin-bottom: 24px; }}
  @media(max-width:900px){{ .cards{{ grid-template-columns:repeat(3,1fr); }} }}
  .card {{ background: var(--card); border-radius: 12px; padding: 20px 16px;
           border: 1px solid var(--border); text-align: center;
           box-shadow: 0 1px 3px rgba(0,0,0,.06); }}
  .card .num {{ font-size: 36px; font-weight: 800; line-height: 1; margin-bottom: 4px; }}
  .card .label {{ font-size: 12px; color: var(--muted); text-transform: uppercase; letter-spacing: .5px; }}
  .card.pass .num {{ color: var(--pass); }}
  .card.fail .num {{ color: var(--fail); }}
  .card.skip .num {{ color: var(--skip); }}
  .card.rate .num {{ color: #2b6cb0; }}
  /* Charts row */
  .charts {{ display: grid; grid-template-columns: 260px 1fr; gap: 20px; margin-bottom: 24px; }}
  @media(max-width:700px){{ .charts{{ grid-template-columns:1fr; }} }}
  .chart-card {{ background: var(--card); border-radius: 12px; padding: 20px;
                 border: 1px solid var(--border); box-shadow: 0 1px 3px rgba(0,0,0,.06); }}
  .chart-card h3 {{ font-size: 14px; font-weight: 600; margin-bottom: 16px; color: var(--muted); }}
  /* Table */
  .section {{ background: var(--card); border-radius: 12px; border: 1px solid var(--border);
              box-shadow: 0 1px 3px rgba(0,0,0,.06); margin-bottom: 24px; overflow: hidden; }}
  .section-header {{ padding: 16px 20px; border-bottom: 1px solid var(--border);
                     display: flex; justify-content: space-between; align-items: center; }}
  .section-header h2 {{ font-size: 15px; font-weight: 700; }}
  .section-header .group-badge {{ background: #ebf8ff; color: #2b6cb0;
                                   padding: 3px 10px; border-radius: 20px; font-size: 12px; }}
  .table-wrap {{ overflow-x: auto; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #f7fafc; color: var(--muted); font-weight: 600; font-size: 11px;
        text-transform: uppercase; letter-spacing: .5px;
        padding: 10px 12px; text-align: left; border-bottom: 2px solid var(--border); }}
  td {{ padding: 10px 12px; border-bottom: 1px solid var(--border); vertical-align: top; }}
  tr:last-child td {{ border-bottom: none; }}
  tr.row-fail {{ background: #fff5f5; }}
  tr.row-pass:hover {{ background: #f0fff4; }}
  .center {{ text-align: center; }}
  .group-cell {{ color: var(--muted); font-size: 12px; white-space: nowrap; }}
  .badge {{ display: inline-block; padding: 3px 8px; border-radius: 20px;
            font-size: 11px; font-weight: 700; white-space: nowrap; }}
  /* History */
  .history-table th, .history-table td {{ padding: 8px 14px; }}
  .hist-row:hover {{ background: #f0f7ff !important; }}
  .hist-row.active {{ background: #dbeafe !important; }}
  /* Modal */
  .modal {{ display: none; position: fixed; inset: 0; background: rgba(0,0,0,.85);
            z-index: 1000; align-items: center; justify-content: center; cursor: pointer; }}
  .modal.open {{ display: flex; }}
  .modal-inner {{ max-width: 90vw; text-align: center; padding: 16px; }}
  /* Footer */
  .footer {{ text-align: center; padding: 20px; color: var(--muted); font-size: 12px; }}
</style>
</head>
<body>

<div class="header">
  <h1>📱 PDF Reader — Auto Test Report</h1>
  <div class="meta" id="header-meta">
    🕐 {run_display} &nbsp;|&nbsp; 📂 {_escape(run_src)}{log_dl_btn}
  </div>
</div>

<div class="container">

  <!-- Run banner (hiện khi đang xem history) -->
  <div class="run-banner" id="run-banner" style="display:none"></div>

  <!-- Summary Cards -->
  <div class="cards">
    <div class="card">
      <div class="num" id="card-total">{total}</div>
      <div class="label">Total TC</div>
    </div>
    <div class="card pass">
      <div class="num" id="card-pass">{passed}</div>
      <div class="label">✅ Passed</div>
    </div>
    <div class="card fail">
      <div class="num" id="card-fail">{failed}</div>
      <div class="label">❌ Failed</div>
    </div>
    <div class="card skip">
      <div class="num" id="card-skip">{skipped}</div>
      <div class="label">⏭ Skipped</div>
    </div>
    <div class="card rate">
      <div class="num" id="card-rate">{rate}</div>
      <div class="label">Pass Rate</div>
    </div>
  </div>

  <!-- Charts + History -->
  <div class="charts">
    <div class="chart-card">
      <h3>Tổng quan</h3>
      <div style="height:220px;position:relative"><canvas id="donutChart"></canvas></div>
    </div>
    <div class="chart-card" style="overflow:auto">
      {_history_section(history) if history
        else '<p style="color:var(--muted);font-size:13px">Chưa có lịch sử chạy trước.</p>'}
    </div>
  </div>

  <!-- TC Tables (updated in-place by JS) -->
  <div id="tc-container">
    {tc_sections_html}
  </div>

</div>

<div class="footer">
  Generated by PDF Reader Auto Test &nbsp;·&nbsp; {run_display}
</div>

<!-- Shared screenshot modal -->
<div id="ss-modal" class="modal" onclick="this.style.display='none'">
  <div class="modal-inner">
    <p id="ss-modal-title" style="color:#fff;margin-bottom:10px;font-weight:bold;font-size:15px"></p>
    <img id="ss-modal-img" src="" style="max-width:90vw;max-height:80vh;border-radius:8px">
    <p style="color:rgba(255,255,255,.5);font-size:12px;margin-top:8px">Click hoặc nhấn Esc để đóng</p>
  </div>
</div>

<script>
// ── Data ──────────────────────────────────────────────────────────────────────
const RUNS_DATA = {runs_data_json};

// ── Chart ─────────────────────────────────────────────────────────────────────
let donutChart = new Chart(document.getElementById('donutChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['Pass', 'Fail', 'Skip / Not Run'],
    datasets: [{{
      data: [{passed}, {failed}, {skipped + not_run}],
      backgroundColor: ['#48BB78','#FC8181','#F6E05E'],
      borderWidth: 2, borderColor: '#fff',
    }}]
  }},
  options: {{
    cutout: '65%',
    plugins: {{
      legend: {{ position: 'bottom', labels: {{ font: {{ size: 12 }} }} }},
      tooltip: {{ callbacks: {{ label: ctx => ` ${{ctx.label}}: ${{ctx.raw}}` }} }},
    }}
  }}
}});

// ── Load run in-place ─────────────────────────────────────────────────────────
let _activeKey = 'current';

function loadRun(key) {{
  const run = RUNS_DATA[key];
  if (!run) return;
  _activeKey = key;

  // Cards
  document.getElementById('card-total').textContent = run.total;
  document.getElementById('card-pass').textContent  = run.passed;
  document.getElementById('card-fail').textContent  = run.failed;
  document.getElementById('card-skip').textContent  = run.skip;
  document.getElementById('card-rate').textContent  = run.rate;

  // Donut chart
  donutChart.data.datasets[0].data = [
    run.passed, run.failed, run.skip + run.not_run
  ];
  donutChart.update();

  // TC tables
  document.getElementById('tc-container').innerHTML = _buildTcContainer(run);

  // Run banner
  const banner = document.getElementById('run-banner');
  if (key === 'current') {{
    banner.style.display = 'none';
  }} else {{
    banner.style.display = 'flex';
    banner.innerHTML = `⏱ Đang xem: <strong>${{run.ts}}</strong> &nbsp; <a href="#" onclick="loadRun('current');return false">← Quay lại hiện tại</a>`;
  }}

  // Highlight history row
  document.querySelectorAll('.hist-row').forEach(r => r.classList.remove('active'));
  if (key !== 'current') {{
    const r = document.getElementById('hist-row-' + key);
    if (r) r.classList.add('active');
  }}
}}

// ── Build TC table HTML from JSON ─────────────────────────────────────────────
const _STATUS_INFO = {{
  'PASS':    {{ bg:'#C6EFCE', fg:'#276221', icon:'✅' }},
  'FAIL':    {{ bg:'#FFC7CE', fg:'#9C0006', icon:'❌' }},
  'SKIP':    {{ bg:'#FFEB9C', fg:'#7D6608', icon:'⏭' }},
  'NOT RUN': {{ bg:'#F2F2F2', fg:'#555555', icon:'⬜' }},
}};

function _buildTcContainer(run) {{
  const groups = {{}};
  for (const c of run.cases) {{
    if (c.status === 'NOT RUN') continue;
    const g = c.group || 'Uncategorized';
    if (!groups[g]) groups[g] = [];
    groups[g].push(c);
  }}
  return Object.entries(groups).map(([g, cs]) => _buildGroupSection(g, cs)).join('');
}}

function _buildGroupSection(groupName, cases) {{
  const total  = cases.length;
  const passed = cases.filter(c => c.status === 'PASS').length;
  const rows   = cases.map(c => _buildTcRow(c)).join('');
  return `<div class="section">
    <div class="section-header">
      <h2>🗂 ${{_esc(groupName)}}</h2>
      <span class="group-badge">${{passed}}/${{total}} Pass</span>
    </div>
    <div class="table-wrap">
      <table>
        <thead><tr>
          <th>TC ID</th><th>Nhóm</th><th>Nội dung</th>
          <th>Các bước</th><th>Kết quả mong đợi</th><th>Kết quả thực tế</th>
          <th>Trạng thái</th><th>Thời gian</th><th>Screenshot</th><th>Video</th><th>Log</th>
        </tr></thead>
        <tbody>${{rows}}</tbody>
      </table>
    </div>
  </div>`;
}}

function _buildTcRow(c) {{
  const st = _STATUS_INFO[c.status] || {{ bg:'#eee', fg:'#333', icon:'?' }};
  const rowCls = c.status === 'FAIL' ? 'row-fail' : c.status === 'PASS' ? 'row-pass' : '';
  const badge  = `<span class="badge" style="background:${{st.bg}};color:${{st.fg}}">${{st.icon}} ${{_esc(c.status)}}</span>`;
  const actClr = c.status==='PASS' ? '#276221' : c.status==='FAIL' ? '#9C0006' : 'inherit';

  const ssCells = (c.screenshots && c.screenshots.length)
    ? c.screenshots.map((b64, i) =>
        `<img src="data:image/png;base64,${{b64}}"
              style="height:44px;border-radius:4px;border:1px solid #ddd;cursor:zoom-in;margin:2px"
              onclick="showSS('${{_esc(c.id)}}',this.src)"
              title="Screenshot ${{i+1}}">`
      ).join('')
    : '—';

  const vidCells = (c.videos && c.videos.length)
    ? c.videos.map(v =>
        `<a href="${{v}}" target="_blank" style="font-size:20px;margin:2px" title="Xem video">🎬</a>`
      ).join('')
    : '—';

  const logCell = c.log
    ? `<details><summary style="cursor:pointer;font-size:12px;color:#2b6cb0;white-space:nowrap">📋 Log</summary>
       <pre style="font-size:11px;white-space:pre-wrap;max-height:300px;overflow:auto;
                   background:#1e1e1e;color:#d4d4d4;padding:8px;border-radius:4px;
                   margin-top:4px;text-align:left">${{_esc(c.log)}}</pre></details>`
    : '—';

  return `<tr class="${{rowCls}}">
    <td><strong>${{_esc(c.id)}}</strong></td>
    <td class="group-cell">${{_esc(c.group)}}</td>
    <td>${{_nl2br(c.title)}}</td>
    <td style="min-width:180px">${{_nl2br(c.steps)}}</td>
    <td style="min-width:180px">${{_nl2br(c.expected)}}</td>
    <td style="min-width:180px;color:${{actClr}}">${{_nl2br(c.actual) || '—'}}</td>
    <td class="center">${{badge}}</td>
    <td class="center" style="white-space:nowrap">${{_esc(c.duration)}}</td>
    <td class="center">${{ssCells}}</td>
    <td class="center">${{vidCells}}</td>
    <td class="center">${{logCell}}</td>
  </tr>`;
}}

// ── Screenshot modal ──────────────────────────────────────────────────────────
function showSS(title, src) {{
  document.getElementById('ss-modal-title').textContent = title + ' — Screenshot';
  document.getElementById('ss-modal-img').src = src;
  document.getElementById('ss-modal').style.display = 'flex';
}}

document.addEventListener('keydown', e => {{
  if (e.key === 'Escape') document.getElementById('ss-modal').style.display = 'none';
}});

// ── Utils ─────────────────────────────────────────────────────────────────────
function _esc(s) {{
  return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;')
                      .replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}}
function _nl2br(s) {{ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;')
                                          .replace(/>/g,'&gt;').replace(/\\n/g,'<br>'); }}
</script>
</body>
</html>"""


# ─── CLI ──────────────────────────────────────────────────────────────────────

def generate(source_xlsx: str = TC_EXCEL, output: str = None,
             run_ts: str | None = None) -> str:
    """Tạo HTML report từ file Excel. Trả về đường dẫn file HTML đã tạo."""
    cases   = _load_tc_excel(source_xlsx)
    history = _load_history(current_ts=run_ts)

    run_info = {
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source":    os.path.basename(source_xlsx),
    }

    html = generate_html(cases, run_info, history, run_ts=run_ts)

    if not output:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        os.makedirs(REPORTS_DIR, exist_ok=True)
        output = os.path.join(REPORTS_DIR, f"dashboard_{ts}.html")

    with open(output, "w", encoding="utf-8") as f:
        f.write(html)

    return output


def main():
    parser = argparse.ArgumentParser(description="Tạo HTML dashboard report")
    parser.add_argument("source", nargs="?", default=TC_EXCEL,
                        help="File Excel chứa kết quả (mặc định: test_cases.xlsx)")
    parser.add_argument("-o", "--output", default=None,
                        help="Đường dẫn file HTML output")
    args = parser.parse_args()

    out = generate(args.source, args.output)
    print(f"[REPORT] HTML dashboard: {out}")
    try:
        import webbrowser
        webbrowser.open(f"file://{os.path.abspath(out)}")
    except Exception:
        pass


if __name__ == "__main__":
    main()
