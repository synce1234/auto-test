#!/usr/bin/env python3
"""
Import test cases từ file Excel gốc (format của QA team) vào test_cases.xlsx.
Chạy: python3 test_cases/import_tc.py <đường_dẫn_file_excel>
"""
import sys
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "test_cases.xlsx")

COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ALT_ROW = "DEEAF1"
COLOR_SECTION_BG = "BDD7EE"

STATUS_COLORS = {
    "PASS":    "C6EFCE",
    "FAIL":    "FFC7CE",
    "SKIP":    "FFEB9C",
    "NOT RUN": "F2F2F2",
}


def parse_source_excel(source_path: str) -> list[dict]:
    """
    Đọc file Excel từ QA team.
    Xử lý merged cells (Phân cấp, Test Title có thể để trống ở các row con).
    """
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active

    rows = []
    last_priority = ""
    last_title = ""
    tc_counter = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Bỏ qua header hoặc row trống
        if not any(cell for cell in row if cell):
            continue

        raw_id, priority, title, precond, steps, expected = row[:6]

        # Bỏ qua nếu là formula string
        if isinstance(raw_id, str) and raw_id.startswith("="):
            raw_id = None

        # Tạo TC_ID chuẩn
        tc_id = f"TC_{tc_counter:03d}"
        tc_counter += 1

        # Kế thừa merged cells
        if priority:
            last_priority = str(priority).strip()
        if title:
            last_title = str(title).strip()

        rows.append({
            "id":           tc_id,
            "priority":     last_priority,
            "title":        last_title,
            "precondition": str(precond or "").strip(),
            "steps":        str(steps or "").strip(),
            "expected":     str(expected or "").strip(),
        })

    return rows


def write_to_template(cases: list[dict], output_path: str = TEMPLATE_PATH):
    """Ghi test cases vào file template xlsx."""

    # Load template hoặc tạo mới
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Test Cases"]
        # Xóa data cũ (giữ header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
    else:
        from create_template import create_template
        create_template()
        wb = openpyxl.load_workbook(output_path)
        ws = wb["Test Cases"]

    col_widths = [15, 18, 40, 40, 50, 40, 40, 14, 14, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    last_priority = None
    for row_idx, tc in enumerate(cases, start=2):
        row_data = [
            tc["id"],
            tc["priority"],
            tc["title"],
            tc["precondition"],
            tc["steps"],
            tc["expected"],
            "",           # actual
            "NOT RUN",    # status
            "",           # duration
            "",           # notes
        ]

        # Màu nền: đổi màu theo nhóm priority
        if tc["priority"] != last_priority:
            last_priority = tc["priority"]
        fill_color = "DEEAF1" if row_idx % 2 == 0 else "FFFFFF"

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if col_idx <= 6:  # chỉ tô màu cột nội dung
                cell.fill = PatternFill("solid", fgColor=fill_color)

        ws.row_dimensions[row_idx].height = 70

    # Bold TC ID
    for row in ws.iter_rows(min_row=2, max_col=1):
        for cell in row:
            if cell.value:
                cell.font = Font(bold=True)

    wb.save(output_path)
    print(f"\n✅ Đã import {len(cases)} test cases → {output_path}")
    for tc in cases:
        print(f"   {tc['id']:8s} | {tc['priority']:15s} | {tc['title'][:50]}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 import_tc.py <source_excel.xlsx>")
        sys.exit(1)

    source = sys.argv[1]
    if not os.path.exists(source):
        print(f"[ERROR] Không tìm thấy file: {source}")
        sys.exit(1)

    print(f"[IMPORT] Đọc từ: {source}")
    cases = parse_source_excel(source)
    print(f"[IMPORT] Tìm thấy {len(cases)} test cases")
    write_to_template(cases)


if __name__ == "__main__":
    main()
